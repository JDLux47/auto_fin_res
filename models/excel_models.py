import io
import json
import sys
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models.nomenclature import nomenclature_list


class ExcelModel:
    def __init__(self):
        self.file_paths = []
        self.dataframes = []
        self.is_valid = []

        self.fot_tax_pct = 0.125
        self.revenue_tax_pct = 0.045
        self.fixed_costs = 40000.0

        # Переменные для логики с руководителем
        self.negative_revenue = 0
        self.director_row = 0

    def add_file(self, file_path: str, index: int) -> bool:
        """Универсальное чтение Excel"""
        while len(self.dataframes) <= index:
            self.dataframes.append(None)
            self.file_paths.append("")
            self.is_valid.append(False)

        if not self._is_excel_file(file_path):
            self.is_valid[index] = False
            return False

        try:
            df = pd.read_excel(file_path, engine=None)
            df = df.dropna(how='all').dropna(axis=1, how='all')

            # Заменяем по индексу
            self.dataframes[index] = df
            self.file_paths[index] = file_path
            self.is_valid[index] = True

            print(f"Файл {index + 1}: {len(df)} строк")
            return True

        except Exception as e:
            print(f"Файл {index + 1}: {e}")
            self.is_valid[index] = False
            return False

    def clear_files(self):
        self.dataframes.clear()
        self.file_paths.clear()
        self.is_valid.clear()

    def _is_excel_file(self, file_path: str) -> bool:
        ext = os.path.splitext(file_path.lower())[1]
        return ext in ['.xlsx', '.xls', '.xlsm']

    def _is_valid_fio(self, text: str) -> bool:
        """ровно 3 слова, каждое начинается с большой буквы"""
        words = text.strip().split()

        # Ровно 3 слова
        if len(words) != 3:
            return False

        # Каждое слово: Заглавная + маленькие (Title Case)
        for word in words:
            if not word or not (
                    word[0].isupper() and  # Первая буква большая
                    len(word) > 1 and  # Более 1 символа
                    all(c.islower() for c in word[1:])  # Остальные маленькие
            ):
                return False

        return True

    def _is_category(self, text: str) -> bool:
        """
        Проверка на категорию
        """
        text_clean = text.strip()
        return text_clean in nomenclature_list

    def all_files_valid(self) -> bool:
        """Все ли 3 файла Excel валидные"""
        return len(self.is_valid) == 3 and all(self.is_valid)

    def get_valid_count(self) -> int:
        """Считает валидные Excel файлы"""
        return sum(1 for valid in self.is_valid if valid)

    def get_employees(self, file_index: int = 0) -> list:
        """
        Извлекает сотрудников и их ЗП
        """
        df = self.dataframes[file_index]

        employees = []

        # Фиксированные столбцы
        name_col = df.columns[0]
        salary_col = df.columns[3]

        for idx, row in df.iterrows():
            name_raw = str(row[name_col]).strip()

            # проверка ФИО ровно 3 слова, каждое с большой буквы
            if self._is_valid_fio(name_raw):
                salary = self._parse_salary(row[salary_col])

                employees.append({
                    'name': name_raw,
                    'salary': salary
                })

        print(f"Найдено {len(employees)} сотрудников (3 слова, большая буква)")
        return employees

    def get_specialists(self, file_index: int = 2) -> list:
        """
        Извлекает спецов по ФИО и сумму их реализации
        """
        df = self.dataframes[file_index].copy()

        specialists = []

        for idx in range(len(df) - 1):
            name_raw = str(df.iloc[idx, 0]).strip()
            next_row_idx = idx + 1
            next_row_col5 = df.iloc[next_row_idx, 4] if next_row_idx < len(df) else None

            is_next_col5_empty = (pd.isna(next_row_col5) or str(next_row_col5).strip() == '')

            # ФИО 3 слова с большой буквы
            if name_raw and len(name_raw.split()) >= 2 and is_next_col5_empty and self._is_valid_fio(name_raw):
                sum_raw = self._parse_salary(df.iloc[idx, 10])

                specialists.append({
                    'name': name_raw,
                    'sum': sum_raw if pd.notna(sum_raw) else 0,
                })

        print(f"Файл {file_index + 1}: Найдено {len(specialists)} сотрудников с услугами")
        return specialists

    def get_managers(self, file_index: int = 1) -> list:
        """
        Извлекает менеджеров по ФИО, услуги, себестоимость и стоимость без ндс
        """
        df = self.dataframes[file_index].copy()

        name_col = df.columns[0]
        price_col = df.columns[2]
        cost_col = df.columns[3]

        managers = []
        current_manager = None

        for idx, row in df.iterrows():
            name_raw = str(row[name_col]).strip()
            if pd.isna(name_raw) or not name_raw:
                continue

            price = self._parse_salary(row[price_col]) or 0
            cost = self._parse_salary(row[cost_col]) or 0

            if self._is_valid_fio(name_raw):
                if current_manager:
                    managers.append(current_manager)

                current_manager = {
                    'name': name_raw,
                    'total_price': price if pd.notna(price) else 0,
                    'total_cost': cost if pd.notna(cost) else 0,
                    'categories': []
                }

            elif current_manager and self._is_category(name_raw):

                service = {
                    'name': name_raw,
                    'price': price if pd.notna(price) else 0,
                    'cost': cost if pd.notna(cost) else 0
                }
                current_manager['categories'].append(service)

        # Последний менеджер
        if current_manager:
            managers.append(current_manager)

        print(f"Найдено {len(managers)} менеджеров")
        for m in managers:
            print(f"   {m['name']}: {len(m['categories'])} услуг из списка")
        return managers

    def _parse_salary(self, salary_raw: str) -> float:
        """Парсит ЗП"""
        salary_raw = str(salary_raw).replace(' ', '').replace(',', '.').replace('₽', '')
        try:
            return float(salary_raw)
        except:
            return 0.0

    def create_result(self, employees: list, specialists: list, managers: list) -> list:
        """
        Объединяет списки в один общий
        """
        result_list = []

        self.negative_revenue = 0  # Сумма отрицательной выручки менеджеров

        for manager in managers:
            if manager['total_price'] < 0:
                self.negative_revenue += manager['total_price']

        for emp in employees:
            person = {
                'name': emp['name'],
                'categories': [],
                'sum': 0,
                'cost_price':0,
                'margin':0,
                'salary': emp['salary'],
                'salary_tax': emp['salary'] * self.fot_tax_pct,
                'sum_tax': None,
                'reg_costs': self.fixed_costs,
                'other_costs_margin': None,
                'res_costs': None,
                'profit_month':None,
                'profit_prc':None
            }

            match_spec = next((s for s in specialists if s['name'] == emp['name']), None)
            match_manager = next((s for s in managers if s['name'] == emp['name']), None)

            if match_manager:
                person['sum'] = match_manager['total_price']
                person['cost_price'] = match_manager['total_cost']
                person['categories'] = match_manager['categories']

            if match_spec:
                person['sum'] = match_spec['sum']
                person['cost_price'] = 0
                person['categories'] = []

            result_list.append(person)

            result_list.sort(key=lambda x: (len(x['categories']) == 0, x['name'] == 'Рыжков Артём Сергеевич', x['name']))

        return result_list

    def create_report(self, result_list):
        """Создает Excel с сотрудниками, заголовками и форматированием"""
        output = io.BytesIO()

        wb = Workbook()
        ws = wb.active
        ws.title = "Финансовая модель ДИТ. Отчёт"

        person_rows = []
        row_num = 3

        row_num = self.persons_in_report(ws, result_list, row_num, person_rows)
        row_num = self.results_in_report(ws, row_num, person_rows)
        self.summary_table_in_report(ws, row_num + 2)
        self.director_in_report(ws, row_num + 2)

        # Табличные границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Применяем границы ко всем ячейкам с данными
        for row in ws.iter_rows(min_row=1, max_row=row_num - 1, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border

        # Ширина колонок
        column_widths = {
            'A': 40,  # Сотрудник, Категория - пошире
            'B': 14,  # Выручка, %
            'C': 14,  # Себестоимость, Сумма
            'D': 14,  # Маржа, Итого
            'E': 14,  # ФОТ
            'F': 14,  # Налоги ФОТ
            'G': 14,  # Налоги выручка
            'H': 14,  # Постоянные расходы
            'I': 14,  # Итого затраты
            'J': 14,  # Рентабельность месяца
            'K': 14   # % рентабельности
        }

        # Применяем ширину колонок
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def persons_in_report(self, ws, result_list, row_num, person_rows):
        # Заголовки
        headers = [
            "Сотрудник", "Выручка (реализации)", "Себестоимость", "Маржа",
            "ФОТ", "Налоги на ФОТ %", "Налоги на выручку %", "Постоянные расходы",
            "Итого затраты", "Рент-сть текущего месяца", "% рент-сти"
        ]

        # Названия колонок
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")

        ws.cell(row=2, column=6, value=self.fot_tax_pct).number_format = '0.0%'
        ws.cell(row=2, column=7, value=self.revenue_tax_pct).number_format = '0.0%'
        ws.cell(row=2, column=8, value=self.fixed_costs)

        for person in result_list:
            # Строка ФИО
            fio_row = row_num

            if len(person['categories']) > 0:
                person_rows.append(row_num)

            # Сохраняем строку с руководителем
            if person['name'] == 'Рыжков Артём Сергеевич':
                self.director_row = row_num
                person_rows.append(self.director_row)

            # Данные строки
            data_row = [
                person['name'],
                person.get('sum', 0),
                person.get('cost_price', 0),
                f"=B{row_num}-C{row_num}",
                person.get('salary', 0),
                f"=E{row_num}*$F$2",
                f"=B{row_num}*$G$2",
                "=$H$2",
                f"=SUM(E{row_num}:H{row_num})",
                f"=D{row_num}-I{row_num}",
                f"=IF(B{row_num}=0,0,J{row_num}/B{row_num}*100)"
            ]

            # Заполняем строку ФИО
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=fio_row, column=col, value=value)

                # Бледно-жёлтый для всей строки ФИО
                if col == 1:  # ФИО - жирный
                    cell.font = Font(bold=True, size=11)
                else:  # Числа
                    cell.number_format = '#,##0.0'

                # Бледно-жёлтый фон для всех ячеек строки
                cell.fill = PatternFill(start_color="FFF2CC", fill_type="solid")

            # Categories под ФИО
            if person.get('categories') and len(person['categories']) > 0:
                row_num += 1  # Пустая строка
                for cat in person['categories']:
                    cat_row = row_num

                    # Категория: только первые 4 колонки
                    ws.cell(row=cat_row, column=1, value=f"{cat['name']}")
                    ws.cell(row=cat_row, column=2, value=cat.get('price', 0) or 0)
                    ws.cell(row=cat_row, column=3, value=cat.get('cost', 0) or 0)
                    ws.cell(row=cat_row, column=4, value=(cat.get('price', 0) - cat.get('cost', 0)) or 0)

                    # Форматирование чисел категорий
                    for col in [2, 3, 4]:
                        cell = ws.cell(row=cat_row, column=col)
                        cell.number_format = '#,##0'

                    row_num += 1
            else:
                row_num += 1

        return row_num

    def results_in_report(self, ws, row_num, person_rows):
        # Надпись ИТОГО
        result_row = row_num
        result_cell = ws.cell(row=result_row, column=1, value="ИТОГО")
        result_cell.font = Font(bold=True, size=12)
        result_cell.alignment = Alignment(horizontal="center")
        red_fill = PatternFill(start_color="FF7514", fill_type="solid")

        # Формулы + форматирование
        columns_data = [
            (2, f"=SUM({','.join(f'B{row}' for row in person_rows)})", '#,##0.0""'),  # Выручка
            (3, f"=SUM({','.join(f'C{row}' for row in person_rows)})", '#,##0.0""'),  # Себестоимость
            (4, f"=SUM({','.join(f'D{row}' for row in person_rows)})", '#,##0.0""'),  # Маржа
            (5, f"=SUM(E3:E{result_row - 1})", '#,##0.0""'),  # ФОТ
            (6, f"=SUM(F3:F{result_row - 1})", '#,##0.0""'),  # Налоги ФОТ
            (7, f"=SUM(G3:G{result_row - 1})", '#,##0.0""'),  # Налоги выручка
            (8, f"=SUM(H3:H{result_row - 1})", '#,##0.0""'),  # Постоянные расходы
            (9, f"=SUM(I3:I{result_row - 1})", '#,##0.0""'),  # Итого затраты
            (10, f"=SUM(J3:J{result_row - 1})", '#,##0.0""'),  # Рентабельность
            (11, f"=J{result_row}/B{result_row}", '0.0%')  # % Рентабильности
        ]

        for col, formula, num_format in columns_data:
            cell = ws.cell(row=result_row, column=col, value=formula)
            cell.number_format = num_format
            cell.fill = red_fill  # Красный фон
            cell.alignment = Alignment(horizontal="right")

        result_cell.fill = red_fill

        return row_num

    def summary_table_in_report(self, ws, row_num):
        # Стили заливки
        light_red_fill = PatternFill(start_color='F8E8E8', end_color='F8E8E8', fill_type='solid')
        light_green_fill = PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Заголовки (строка)
        headers = ['Категория', '%', 'Сумма', 'Итого']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row_num += 1

        data = [
            {'category': 'Проектники', 'pct': 0.8, 'base': 0, 'result': f'=C{row_num}*B{row_num}'},
            {'category': 'Внедрение', 'pct': 0.7, 'base': 0, 'result': f'=C{row_num + 1}*B{row_num + 1}'},
            {'category': 'Субподряд', 'pct': 1.0, 'base': self.negative_revenue * -1, 'result': f'=C{row_num + 2}*B{row_num + 2}'},
            {'category': 'Отрицательный взаиморасчёт', 'pct': 1.0, 'base': 0, 'result': f'=C{row_num + 3}*B{row_num + 3}'},
            {'category': 'Реализации из иных отделов', 'pct': 0.7, 'base': 0, 'result': f'=C{row_num + 4}*B{row_num + 4}'},
            {'category': 'Взаиморасчёты', 'pct': 1.0, 'base': 0, 'result': f'=C{row_num + 5}*B{row_num + 5}'},
        ]

        for i, row_data in enumerate(data):
            row = row_num + i
            ws.cell(row=row, column=1, value=row_data['category'])
            ws.cell(row=row, column=2, value=row_data['pct']).number_format = '0%'
            ws.cell(row=row, column=3, value=row_data['base']).number_format = '#,##0.0""'
            ws.cell(row=row, column=4, value=row_data['result']).number_format = '#,##0.0""'

            # Применяем стили
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, size=12) if col == 1 else Font(bold=False, size=12)
                cell.fill = light_red_fill if i < 4 else light_green_fill
                cell.border = thin_border

    def director_in_report(self, ws, row_num):
        ws.cell(row=self.director_row, column=3, value=f'=D{row_num + 1}+D{row_num + 2}+D{row_num + 3}+D{row_num + 4}')
        ws.cell(row=self.director_row, column=2, value=f'=D{row_num + 5}+D{row_num + 6}')
